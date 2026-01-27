from openpyxl import load_workbook, Workbook

wb = load_workbook("data.xlsx")
sheet = wb.active

new_wb = Workbook()
new_sheet = new_wb.active

first = True
kept = 0
for row in sheet.iter_rows(values_only=True):
    if first:
        new_sheet.append(row)
        first = False
        continue

    try:
        value = int(row[column_index])
    except:
        continue
    
    if value >= threshold:
        new_sheet.append(row)
        kept += 1

new_wb.save(output_name)
print(f"Done. {kept} rows saved to {output_name}")