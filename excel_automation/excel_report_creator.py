import os
import sys
from openpyxl import load_workbook, Workbook

SCRIPT_DIR = os.path.dirname(__file__)

def menu():
    print("--- Excel report filter utility ---")
    print("1: Create a new excel report (salary ≥ 4000)")
    print("2: Quit")
    choice = input("> ")

    if choice == "1":
        option_1()
    elif choice == "2":
        option_2()
    else:
        print("Invalid option")

def option_1():
    input_file = input("Excel file name (must be in this folder): ")
    output_name = input("New file name (without .xlsx, IF AN EXCEL FILE WITH THE SAME NAME ALREADY EXISTS IT WILL BE OVERWRITEN, BE AWARE, type n to cancel): ")

    if output_name.lower() == "n":
        print("Returning to the main menu.")
        return

    try:
        input_path = os.path.join(SCRIPT_DIR, input_file)
        output_path = os.path.join(SCRIPT_DIR, f"{output_name}.xlsx")

        wb = load_workbook(input_path)
        sheet = wb.active

        new_wb = Workbook()
        new_sheet = new_wb.active

        first = True
        for row in sheet.iter_rows(values_only=True):
            if first:
                new_sheet.append(row)
                first = False
                continue

            person_name, age, salary = row
            if isinstance(salary, (int, float)) and salary >= 4000:
                new_sheet.append(row)

        new_wb.save(output_path)
        print("New report created successfully!")
        input("Press Enter to return to menu")

    except Exception as e:
        print(f"Error: {e}")

def option_2():
    print("Quitting...")
    sys.exit()

while True:
    menu()