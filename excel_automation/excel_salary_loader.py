import os
import sys
from openpyxl import load_workbook

def menu():
    print("=== Excel salary filter ===")
    print("Choose an option: ")
    print("1- Filter a salary from a excel file")
    print("2 - Quit")

def option_1():
    print("Type the file name or give the whole file path: ")
    file = input("")
    try:
        base_dir = os.path.dirname(__file__)
        file_path = os.path.join(base_dir, file)

        wb = load_workbook(file)
        sheet = wb.active

        first = True

        for row in sheet.iter_rows(values_only= True):
            if first:
                first = False
                continue
            name, _, salary = row
            print(f"{name} earns {salary}")
        input("Press enter to return to the menu")
    except Exception as e:
        print(f"An error ocurred, perhaps you didn't type the file correctly or it doesn't exist. Error message: {e}")
        print("Returning to the menu...")
        return

def option_2():
    print("Quitting...")
    sys.exit()

while True:
    menu()
    choice = input("")
    if choice == "1":
        option_1()
    elif choice == "2":
        option_2()
    else:
        print("Invalid Option")