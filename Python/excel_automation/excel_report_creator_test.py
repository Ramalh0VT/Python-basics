from openpyxl import load_workbook, Workbook

wb = load_workbook("data.xlsx")
sheet = wb.active

new_wb = Workbook()
new_sheet = new_wb.active

first = True

for row in sheet.iter_rows(values_only=True):
    if first:
        new_sheet.append(row)
        first = False
        continue
    name, age, salary = row
    try:
        if salary >= 3500:
            new_sheet.append(row)
    except:
        continue
new_wb.save("high_earners.xlsx")
print("Report created: high_earners.xlsx")